"""
值班表管理模块 — 上传、解析、存储
"""
import os
import re
from datetime import datetime
import openpyxl
from database import get_connection
from models import DutyTableModel
from config import APP_CONFIG


def _fmt_date(val):
    """将 Excel 日期值统一转为 yyyy/mm/dd 格式"""
    if val is None:
        return ''
    # openpyxl 的 datetime 类型
    if isinstance(val, datetime):
        return val.strftime('%Y/%m/%d')
    s = str(val).strip()
    # 尝试匹配常见日期格式
    m = re.match(r'(\d{4})[-/](\d{1,2})[-/](\d{1,2})', s)
    if m:
        return f"{m.group(1)}/{int(m.group(2)):02d}/{int(m.group(3)):02d}"
    return s


def parse_xlsx(file_path, sheet_name=None):
    """解析 xlsx 文件，返回记录列表
    
    Args:
        file_path: Excel 文件路径
        sheet_name: 工作表名称，None 表示使用活动工作表
    """
    wb = openpyxl.load_workbook(file_path, read_only=True)
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    records = []
    for row in rows:
        if row[0] and row[2]:
            records.append({
                'date': _fmt_date(row[0]),
                'weekday': str(row[1]).strip() if row[1] else '',
                'person': str(row[2]).strip(),
            })
    wb.close()
    return records


def parse_xlsx_from_bytes(file_bytes, file_name, sheet_name=None):
    """从字节流解析 xlsx（前端上传场景）"""
    upload_dir = APP_CONFIG['upload_dir']
    os.makedirs(upload_dir, exist_ok=True)
    file_path = os.path.join(upload_dir, file_name)

    with open(file_path, 'wb') as f:
        f.write(file_bytes)

    records = parse_xlsx(file_path, sheet_name=sheet_name)
    return records


def save_duty_table(file_name, records):
    """将解析后的值班记录保存到数据库"""
    if not records:
        return {'success': False, 'error': '未解析到有效数据'}

    name = os.path.splitext(file_name)[0]

    conn = get_connection()
    table_id = DutyTableModel.create(conn, name, file_name, len(records))
    DutyTableModel.insert_records(conn, table_id, records)
    conn.commit()
    conn.close()

    return {
        'success': True,
        'id': table_id,
        'name': name,
        'file_name': file_name,
        'record_count': len(records),
    }


def get_all_tables():
    """获取所有值班表"""
    conn = get_connection()
    tables = DutyTableModel.get_all(conn)
    conn.close()
    return [dict(t) for t in tables]


def get_table_records(table_id):
    """获取某值班表的记录明细"""
    conn = get_connection()
    records = DutyTableModel.get_records(conn, int(table_id))
    conn.close()
    return [dict(r) for r in records]


def get_today_duty_persons(table_id):
    """获取今天的值班人员列表（用于通知消息）"""
    from datetime import date
    today_str = date.today().strftime('%Y-%m-%d')

    conn = get_connection()
    rows = conn.execute(
        'SELECT person FROM duty_records WHERE table_id = ? AND duty_date = ?',
        (int(table_id), today_str)
    ).fetchall()
    conn.close()

    return [r['person'] for r in rows]


def delete_duty_table(table_id):
    """删除值班表及其记录"""
    conn = get_connection()
    DutyTableModel.delete(conn, int(table_id))
    conn.commit()
    conn.close()
    return {'success': True}
