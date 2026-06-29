"""
钉钉自定义机器人消息发送模块
文档：https://open.dingtalk.com/document/orgapp/custom-robot-access
签名算法：timestamp + "\n" + secret → HMAC-SHA256 → Base64 → URL Encode
"""
import requests
import logging
import time
import hmac
import hashlib
import base64
import urllib.parse
from database import get_connection
from models import BotConfigModel, DutyTableModel, SendLogModel
from config import APP_CONFIG

logger = logging.getLogger(__name__)


def _build_duty_markdown(table_id, title=None):
    """根据值班表构建今日值班 + 下次预告的 Markdown 通知"""
    from datetime import date
    conn = get_connection()
    table = DutyTableModel.get_by_id(conn, int(table_id))
    records = DutyTableModel.get_records(conn, int(table_id))
    conn.close()

    heading = title or '值班信息通知'
    if not records:
        return f"## {heading}\n\n暂无值班信息"

    today_str = date.today().strftime('%Y/%m/%d')

    # 查找今日值班记录
    today_idx = None
    for i, r in enumerate(records):
        if r['duty_date'] == today_str:
            today_idx = i
            break

    lines = []

    if today_idx is not None:
        r = records[today_idx]
        lines.append(f"## **今日值班信息：**")
        lines.append(f"{r['duty_date']} {r['weekday']} **{r['person']}**")
    else:
        lines.append(f"## **今日值班信息：**")
        lines.append(f"今日（{today_str}）暂无值班安排")

    # 查找下一次值班（今日之后的第一条）
    next_idx = None
    if today_idx is not None and today_idx + 1 < len(records):
        next_idx = today_idx + 1
    else:
        for i, r in enumerate(records):
            if r['duty_date'] > today_str:
                next_idx = i
                break

    if next_idx is not None:
        r = records[next_idx]
        lines.append("")
        lines.append("---")
        lines.append("")
        lines.append(f"## **下次值班预告：**")
        lines.append(f"{r['duty_date']} {r['weekday']} **{r['person']}**")

    lines.append("")
    lines.append("---")
    lines.append("")
    lines.append("请值班人员准时签到值班！")

    return "\n".join(lines)


def build_duty_markdown_from_excel(file_path, title=None, sheet_name=None):
    """从上传的 Excel 文件构建今日值班 + 下次预告的 Markdown 通知
    格式：第1列=日期，第2列=星期，第3列=人员
    
    Args:
        file_path: Excel 文件路径
        title: 通知标题
        sheet_name: 工作表名称，None 表示使用活动工作表
    """
    import os
    import re
    from datetime import date, datetime as dt
    import openpyxl

    if not os.path.exists(file_path):
        return None

    wb = openpyxl.load_workbook(file_path, read_only=True)
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active
    all_rows_data = list(ws.iter_rows(values_only=True))
    wb.close()

    today_str = date.today().strftime('%Y/%m/%d')
    records = []
    for row_data in all_rows_data[1:]:  # 跳过表头行
        if not row_data[0] or not row_data[2]:
            continue
        date_val = row_data[0]
        if isinstance(date_val, dt):
            date_str = date_val.strftime('%Y/%m/%d')
        else:
            m = re.match(r'(\d{4})[-/](\d{1,2})[-/](\d{1,2})', str(date_val).strip())
            date_str = f"{m.group(1)}/{int(m.group(2)):02d}/{int(m.group(3)):02d}" if m else str(date_val).strip()
        records.append({
            'duty_date': date_str,
            'weekday': str(row_data[1]).strip() if row_data[1] else '',
            'person': str(row_data[2]).strip(),
        })

    heading = title or '值班信息通知'
    if not records:
        return f"## {heading}\n\n暂无值班信息"

    # 查找今日值班记录
    today_idx = None
    for i, r in enumerate(records):
        if r['duty_date'] == today_str:
            today_idx = i
            break

    lines = [f'## {heading}', '']
    if today_idx is not None:
        r = records[today_idx]
        lines.append('<h3> **今日值班信息：**</h3>')    
        lines.append(f"{r['duty_date']} {r['weekday']} **{r['person']}**")
    else:
        lines.append('<h3> **今日值班信息：**</h3>')
        lines.append(f"今日（{today_str}）暂无值班安排")

    # 下一次值班（今天之后的第一条）
    next_idx = None
    if today_idx is not None and today_idx + 1 < len(records):
        next_idx = today_idx + 1
    else:   
        for i, r in enumerate(records):
            if r['duty_date'] > today_str:
                next_idx = i
                break

    if next_idx is not None:
        r = records[next_idx]
        lines.append('')
        #lines.append('---')
        lines.append('')
        lines.append('<h3> **下次值班预告：**</h3>')
        lines.append(f"{r['duty_date']} {r['weekday']} **{r['person']}**")

    lines.append('')
    lines.append('---')
    lines.append('')
    lines.append('请值班人员准时签到值班！')

    return '\n'.join(lines)


def send_duty_notification(bot_ids, table_id, at_all=False, custom_text='', title=None):
    """发送值班通知到指定机器人列表，可选拼接自定义文本"""
    message_title = title or '值班信息通知'
    message = _build_duty_markdown(table_id, title=message_title)
    if custom_text and custom_text.strip():
        message += '\n\n---\n\n' + custom_text.strip()
    conn = get_connection()
    table = DutyTableModel.get_by_id(conn, int(table_id))
    conn.close()
    return _send_to_bots(
        bot_ids=bot_ids,
        title=message_title,
        message=message,
        at_all=at_all,
        log_type='duty',
        table_id=int(table_id) if table_id else None,
    )


def send_custom_message(bot_ids, message_text, at_all=False):
    """发送自定义消息到指定机器人列表"""
    return _send_to_bots(
        bot_ids=bot_ids,
        title="💬 通知消息",
        message=message_text,
        at_all=at_all,
        log_type='custom',
    )


def _build_signed_url(webhook, secret):
    """钉钉加签：为 Webhook URL 附加 timestamp 和 sign 参数"""
    if not secret:
        return webhook

    timestamp = str(round(time.time() * 1000))
    string_to_sign = f"{timestamp}\n{secret}"
    hmac_code = hmac.new(
        secret.encode('utf-8'),
        string_to_sign.encode('utf-8'),
        digestmod=hashlib.sha256
    ).digest()
    sign = urllib.parse.quote_plus(base64.b64encode(hmac_code))

    separator = '&' if '?' in webhook else '?'
    return f"{webhook}{separator}timestamp={timestamp}&sign={sign}"


def _send_to_bots(bot_ids, title, message, at_all, log_type, table_id=None):
    """核心发送逻辑：遍历机器人发送钉钉消息 + 写日志"""
    conn = get_connection()
    bots = BotConfigModel.get_all(conn)
    bot_map = {b['id']: b for b in bots}

    bot_names = []
    all_success = True
    last_error = ''

    for bot_id in bot_ids:
        bot = bot_map.get(int(bot_id))
        if not bot:
            logger.warning(f"机器人 ID={bot_id} 不存在，跳过")
            continue

        bot_names.append(bot['name'])

        # 截断过长消息
        max_len = APP_CONFIG['max_message_length']
        text = message
        if len(text) > max_len:
            text = text[:max_len - 30] + "\n\n...（消息过长已截断）"

        # 构建钉钉 Markdown 消息体
        payload = {
            "msgtype": "markdown",
            "markdown": {
                "title": title,
                "text": text,
            },
        }

        # 钉钉 @所有人 通过 at 字段实现
        if at_all:
            payload["at"] = {
                "isAtAll": True,
            }

        try:
            signed_url = _build_signed_url(bot['webhook'], bot['secret'])
            resp = requests.post(
                signed_url,
                json=payload,
                timeout=APP_CONFIG['webhook_timeout'],
            )
            result = resp.json()
            # 兼容两种响应格式：标准钉钉 {"errcode":0} / 内网 {"success":true}
            is_success = (
                resp.status_code == 200 and
                (result.get('errcode') == 0 or result.get('success') is True)
            )
            if not is_success:
                all_success = False
                last_error = result.get('errmsg', str(result))[:200]
                logger.error(f"发送到 {bot['name']} 失败: {last_error}")
            else:
                logger.info(f"发送到 {bot['name']} 成功")
        except requests.RequestException as e:
            all_success = False
            last_error = str(e)[:200]
            logger.error(f"发送到 {bot['name']} 异常: {last_error}")

    if not bot_names:
        conn.close()
        return {'success': False, 'error': '没有有效的机器人'}

    # 获取表名（仅 duty 类型）
    table_name = ''
    if table_id:
        table = DutyTableModel.get_by_id(conn, table_id)
        table_name = table['name'] if table else ''

    # 记录发送日志
    summary = message[:100].replace('\n', ' ')
    SendLogModel.create(conn, {
        'log_type': log_type,
        'bot_names': '、'.join(bot_names),
        'message_summary': summary,
        'table_name': table_name,
        'at_all': 1 if at_all else 0,
        'status': 'success' if all_success else 'failed',
        'error_message': last_error,
    })
    conn.commit()
    conn.close()

    return {
        'success': all_success,
        'bot_names': '、'.join(bot_names),
        'error': last_error if not all_success else '',
    }
